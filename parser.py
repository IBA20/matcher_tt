import contextlib
import datetime
import os
import sqlite3

import openpyxl
from dotenv import load_dotenv


class FileFormatError(Exception):
    pass


def validate_row(
        row,
        type_mask=(
            (int, float), str, (int, float), (int, float), (int, float),
            (int, float), (int, float), (int, float), (int, float), (int, float)
        )
):
    if not all(
        [
            isinstance(value, value_type)
            for value, value_type in zip(row, type_mask)
        ]
    ):
        raise FileFormatError


def parse_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    parsed_data = []
    for row in worksheet.iter_rows(4, worksheet.max_row, values_only=True):
        validate_row(row)
        parsed_data.append(list(row))
    return parsed_data


def create_db(cursor):
    cursor.execute(
        '''
        CREATE TABLE parsed_data(
            row_id INT, 
            date DATE, 
            company VARCHAR(50), 
            fact_qliq_data1 INT, 
            fact_qliq_data2 INT, 
            fact_qoil_data1 INT, 
            fact_qoil_data2 INT, 
            forecast_qliq_data1 INT, 
            forecast_qliq_data2 INT, 
            forecast_qoil_data1 INT, 
            forecast_qoil_data2 INT
        )
        '''
    )


def add_random_dates(data):
    start_date = datetime.date(2023, 3, 1)
    rows_per_date = max(len(data) // 30 + 1, 2)
    for n, row in enumerate(data):
        date = start_date + datetime.timedelta(days=n // rows_per_date)
        row.insert(1, date)
    return data


def get_table_with_subtotals(cursor):
    query = cursor.execute(
        '''
            SELECT * FROM
            (SELECT * FROM parsed_data
            UNION
            SELECT "", date, "SUBTOTAL", SUM(fact_qliq_data1), SUM(fact_qliq_data2), 
            SUM(fact_qoil_data1), SUM(fact_qoil_data2), SUM(forecast_qliq_data1), 
            SUM(forecast_qliq_data2), SUM(forecast_qoil_data1), 
            SUM(forecast_qoil_data2) FROM parsed_data GROUP BY date          
            ORDER BY date)
            UNION ALL
            SELECT "", "", "GRAND TOTAL", SUM(fact_qliq_data1), SUM(fact_qliq_data2), 
            SUM(fact_qoil_data1), SUM(fact_qoil_data2), SUM(forecast_qliq_data1), 
            SUM(forecast_qliq_data2), SUM(forecast_qoil_data1), 
            SUM(forecast_qoil_data2) FROM parsed_data;
            '''
    )
    return query.fetchall()


def print_table(table):
    print(f'{"":^28}{"fact":^24}{"forecast":^24}')
    print(f'{"" :^28}{"Qliq":^12}{"Qoil":^12}{"Qliq":^12}{"Qoil":^12}')
    print(f'{"id":^4}{"date":^12}{"company":^12}{"data1":>6}{"data2":>6}{"data1":>6}'
          f'{"data2":>6}{"data1":>6}{"data2":>6}{"data1":>6}{"data2":>6}')
    for row in table:
        print(f'{row[0]:>4}{row[1]:^12}{row[2]:<12}', end='')
        for value in row[3:]:
            print(f'{value:>6}', end='')
        print()


def main():
    load_dotenv()
    db_path = os.getenv('DB_PATH', 'db.sqlite3')
    raw_data_path = os.getenv('RAW_DATA_PATH', 'raw_data/')
    parsed_data = []
    for file in os.listdir(raw_data_path):
        filename, extension = os.path.splitext(file)
        if extension != '.xlsx' or filename.endswith('_processed'):
            continue
        try:
            parsed_data.extend(parse_excel(os.path.join(raw_data_path, file)))
            os.rename(
                os.path.join(raw_data_path, file),
                os.path.join(raw_data_path, f'{filename}_processed.xlsx'))
        except FileFormatError:
            print(f'Неверный формат данных в файле {file}. Файл не обработан.')

    parsed_data_with_dates = add_random_dates(parsed_data)

    con = sqlite3.connect(db_path)
    cur = con.cursor()
    with contextlib.suppress(sqlite3.OperationalError):
        create_db((cur))

    cur.executemany(
        "INSERT INTO parsed_data VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        parsed_data_with_dates
    )
    con.commit()
    table_with_subtotals = get_table_with_subtotals(cur)
    con.close()
    print_table(table_with_subtotals)


if __name__ == '__main__':
    main()
